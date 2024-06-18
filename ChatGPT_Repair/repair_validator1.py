import os
import time
import openai
import json
import openpyxl


def is_row_empty(row):
    for cell in row:
        if cell.value:
            return False
    return True


def analysis_repair(path,output,a,to_replace):

    # workbook = openpyxl.load_workbook(output)
    workbook = openpyxl.Workbook()
    print("open "+output)
    sheet = workbook.active
    sheet["A1"] = "file name"
    sheet["B1"] = "b1"
    sheet["C1"] = "b2"
    sheet["D1"] = "b3"
    sheet["E1"] = "b4"
    sheet["F1"] = ""
    sheet["G1"] = "r1"
    sheet["H1"] = "r2"
    sheet["I1"] = "r3"
    sheet["J1"] = "r4"
    sheet["K1"] = ""
    sheet["L1"] = "re1"
    sheet["M1"] = "re2"
    sheet["N1"] = "re3"
    sheet["O1"] = "re4"
    sheet["P1"] = ""
    sheet["Q1"] = "b1=r1&b4=r4"
    sheet["R1"] = "selected id"
    sheet["S1"] = "consistent with selection"
    sheet["T1"] = "correct matching"
    sheet["U1"] = "repaired correct"
    sheet["V1"] = "multi repaired stm"
    sheet["W1"] = "multi stm correct"
    print("max line: " + str(sheet.max_row))
    i = 2
    for root, dirs, files in os.walk(path):
        for filename in files:
            print(filename)
            # print(a)
            # print(os.path.join(root, filename))
            # if str(a) in filename and filename.endswith(".txt") and "_all" not in filename and "EditProjectTest_36" in filename:
            if filename.endswith(".txt"):
                filepath = os.path.join(root, filename)
                str_arr = filename.split("_")
                for j in range(2,200):
                    row_to_check=sheet[j]
                    if is_row_empty(row_to_check):
                        i=j
                        break

                name = str_arr[0] + "_" + str_arr[1]
                # sheet.cell(row=1, column=i,value=name)
                broken_filename = os.path.join(root.replace(to_replace, "broken_statement"), name+".txt")
                print(broken_filename)
                if "Login_12" in broken_filename:
                    broken_filename=broken_filename.replace("Login_12.txt","Login_12_TestAddAdmin.txt")
                if "Category_8" in broken_filename:
                    broken_filename=broken_filename.replace("Category_8","Category_8_AssignCategoryToUser")
                if "Course_51" in broken_filename:
                    broken_filename=broken_filename.replace("Course_51","Course_51_TestRemoveCourse")
                with open(broken_filename,'r') as f:
                    data0=f.read()
                min_index1 = find_index3(data0, "Assert")
                min_index2 = find_index3(data0, "driver")
                min_index = min(min_index1, min_index2)
                min_index3 = find_index3(data0, "assert")
                min_index = min(min_index3, min_index)

                sub_str1 = data0[min_index:]
                min_index3 = find_index3(sub_str1, ");")
                sub_str2 = sub_str1[:min_index3 + 2]
                # print(sub_str2)
                b1,b2,b3,b4,b_locator_index=seperateStm(sub_str2)

                # print(filepath)
                with open(filepath, 'r', encoding="utf-8") as f:
                    data = f.read()
                    # print(data)
                # extract answer id, return 10000 means not found
                min_index1=find_index3(data,"Assert")
                min_index2=find_index3(data,"driver")
                min_index=min(min_index1,min_index2)
                min_index3=find_index3(data,"assert")
                min_index=min(min_index3,min_index)

                sub_str1 = data[min_index:]
                min_index3=find_index3(sub_str1,");")
                sub_str2 =sub_str1[:min_index3+2]
                # print(sub_str2)
                r1,r2,r3,r4,r_locator_index=seperateStm(sub_str2)

                if r_locator_index==-1:
                    print("selection is same as before, no new repair")
                    continue
                re1=0
                re2=0
                re3=0
                re4=0
                if b1.replace("\n","")==r1.replace("\n",""):
                    re1=1
                if b2.replace("\n","")==r2.replace("\n",""):
                    re2=1
                if b4.replace("\n","")==r4.replace("\n",""):
                    re4=1

                sheet.cell(row=i,column=1,value=name)
                sheet.cell(i,2,b1)
                sheet.cell(i, 3, b2)
                sheet.cell(i, 4, b3)
                sheet.cell(i, 5, b4)

                sheet.cell(i, 7, r1)
                sheet.cell(i, 8, r2)
                sheet.cell(i, 9, r3)
                sheet.cell(i, 10, r4)

                sheet.cell(i, 12, re1)
                sheet.cell(i, 13, re2)
                sheet.cell(i, 14, re3)
                sheet.cell(i, 15, re4)


    workbook.save(output)
                # print("save "+output)
                # i+=1


def seperateStm(stm):
    print("current repaired statement:")
    print(stm)
    if stm.find("driver.findElement(")==-1:
        return '','','','',-1

    index1 = stm.find("driver.findElement(")+len("driver.findElement(")
    # print(index1)
    sub0 = stm[0:index1]
    sub1 = stm[index1:]
    count=0
    index2=-1
    locator_index=0

    if ".click" in sub1:
        index2 = sub1.index(".click")
    elif ".sendKey" in sub1:
        index2 = sub1.index(".sendKey")
    elif ".clear" in sub1:
        index2 = sub1.index(".clear")
    elif ".getText" in sub1:
        index2 = sub1.index(".getText")
    elif ".getAttribute" in sub1:
        index2 = sub1.index(".getAttribute")

    sub2=sub1[:index2]
    locator="xpath"
    # index3=-1
    print(sub2)

    if "By.xpath" in sub2:
        index3 = sub2.rindex("By.xpath")+len("By.xpath")
        locator = "xpath"
        locator_index = 9
    elif "By.name" in sub2:
        index3 = sub2.rindex("By.name")+len("By.name")
        locator = "name"
        locator_index = 4
    elif "By.linkText" in sub2:
        index3 = sub2.rindex("By.linkText")+len("By.linkText")
        locator = "linkText"
        locator_index = 7
    elif "By.tagName" in sub2:
        index3 = sub2.rindex("By.tagName")+len("By.tagName")
        locator = "tagName"
        locator_index = 6
    elif "By.id" in sub2:
        index3 = sub2.rindex("By.id")+len("By.id")
        locator = "id"
        locator_index = 3
    elif "By.cssSelector" in sub2:
        index3 = sub2.rindex("By.cssSelector")+len("By.cssSelector")
        locator = "cssSelector"
    elif "By.className" in sub2:
        index3 = sub2.rindex("By.className")+len("By.className")
        locator = "className"

    sub3 = sub2[index3:]
    sub3 = sub3[2:len(sub3) - 3]
    index4 = sub1.rindex(sub3)+len(sub3)
    sub4 = sub1[index4:]
    print(sub0+"   "+locator+"   "+sub3+"   "+sub4)
    # print()
    return sub0,locator,sub3,sub4,locator_index

def check_locator(locator1,content1,locator2,content2):
    same_locator=False
    print("----")
    relative_xpath1=False
    relative_xpath2=False
    if locator1!="xpath" and locator1==locator2:
        same_locator=True
    if locator1=="xpath" and ("=" in content1 or "contain" in content1):#and "[@"in content1:
        # print(locator1)
        relative_xpath1=True
    if locator2=="xpath" and ("=" in content1 or "contain" in content1):#and "[@" in content2:
        # print(locator2)
        relative_xpath2=True
    if relative_xpath1 and relative_xpath2:
        same_locator=True
        relative_xpath1=True
    return same_locator,relative_xpath1,relative_xpath2


def check_ancestor(xpath1,absolute1,xpath2,absolute2):
    ancestor1=''
    ancestor2=''
    last1=''
    last2=''
    ancestor_same=False
    end=0
    for c in xpath1:
        if (c!="]") and end==0:
            ancestor1+=c
            continue
        elif c==']' and end==0:
            end=1
            continue
        last1+=c
    end=0
    for c in xpath2:
        if (c!="]") and end==0:
            ancestor2+=c
            continue
        elif c==']' and end==0:
            end=1
            continue
        last2+=c
    print(ancestor1+" : "+ancestor2)
    if ancestor1!=ancestor2:
        return False
    print(last1+" :  "+last2)
    print("-----")
    ancestor_xpath=ancestor1.replace(last1,"")# remove the rest xpath
    if ancestor2.replace(ancestor_xpath,"")==last2:
        return True

# def check_update_attr(locator,selected_id,attribute):

def fill_taget_xpath(result,attributes):
    workbook_re = openpyxl.load_workbook(result)
    sheet_re = workbook_re.active

    workbook_attr = openpyxl.load_workbook(attributes)
    sheet_attr = workbook_attr.active

    for y in range(2,sheet_re.max_row):
        filename0 = sheet_re.cell(row=y, column=1).value
        selected_id = sheet_re.cell(row=y, column=18).value
        # print(filename0)
        target_xpath=''
        selected_xpath=''
        for i in range(2, sheet_attr.max_row):
            filename = sheet_attr.cell(row=i, column=1).value
            flag = 0
            position = 0
            if filename == None:
                continue
            if "-" in filename:
                continue

            # print(filename0 + "   :  " + filename)
            if filename == filename0:
                target_xpath = sheet_attr.cell(row=i, column=9).value
                # print(filename)
                # print(i)
                for z in range(1, 11):
                    attr_id=sheet_attr.cell(row=i+z,column=2).value
                    print("selected id: "+str(selected_id)+" : "+str(attr_id))
                    attr_xpath = sheet_attr.cell(row=i + z, column=9).value
                    if int(selected_id)==int(attr_id):
                        selected_xpath=attr_xpath
                        break

        sheet_re.cell(row=y, column=17, value=target_xpath)
        sheet_re.cell(row=y, column=19, value=selected_xpath)

    workbook_re.save(result)


def find_index3(data,str):
    try:
        index=data.index(str)
    except ValueError:
        index=10000
    return index

def find_index(data,str):
    try:
        index=data.index(str)
    except ValueError:
        index=10000
    sub=data[:index]
    if "is" in sub and any(c.isdigit() for c in sub):
        return index
    if "is" in str:
        sub=data[index:]
        if any(c.isdigit() for c in sub):
            return index
    else:
        return 10000
def find_index2(data,str):
    try:
        index=data.rfind(str)
    except ValueError:
        index=10001
    return index-5


if __name__ == "__main__":

    methods=["vista","water","xpath"] # 1102
    # methods=["vista"]
    for method in methods:
        for i in range(0,4):
            answers_path = "...\\chatgpt_answers" + str(i) + "\\"+method
            output_path = ".../"+method+"_repair"+str(i)+".xlsx"
            to_replace = "chatgpt_answers" + str(i) + "\\" + method
            print(to_replace)
            analysis_repair(answers_path, output_path, a=a3, to_replace=to_replace)




