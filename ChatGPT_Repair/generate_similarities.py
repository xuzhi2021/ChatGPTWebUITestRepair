import openpyxl
import math
import os
def points_distance(x1,y1,x2,y2):
    # x1=int(s1.split(",")[0])
    # y1=int(s1.split(",")[1])
    # x2=int(s2.split(",")[0])
    # y2=int(s2.split(",")[1])
    x1=int(x1)
    y1=int(y1)
    x2=int(x2)
    y2=int(y2)
    dx = x2 - x1
    dy = y2 - y1
    dist = math.sqrt(dx*dx + dy*dy)
    return dist

def levenshtein_distance(s, t):
    m = len(s)
    n = len(t)
    d = [[0] * (n + 1) for i in range(m + 1)]
    for i in range(m + 1):
        d[i][0] = i
    for j in range(n + 1):
        d[0][j] = j
    for j in range(1, n + 1):
        for i in range(1, m + 1):
            if s[i - 1] == t[j - 1]:
                cost = 0
            else:
                cost = 1
            d[i][j] = min(d[i - 1][j] + 1, # deletion
                          d[i][j - 1] + 1, # insertion
                          d[i - 1][j - 1] + cost) # substitution
    return d[m][n]
def xpath(s,t):
    m = len(s)
    n = len(t)
    d = [[0] * (n + 1) for i in range(m + 1)]
    for i in range(m + 1):
        d[i][0] = i
    for j in range(n + 1):
        d[0][j] = j
    for j in range(1, n + 1):
        for i in range(1, m + 1):
            if s[i - 1] == t[j - 1]:
                cost = 0
            else:
                cost = 1
            d[i][j] = min(d[i - 1][j] + 1,  # deletion
                          d[i][j - 1] + 1,  # insertion
                          d[i - 1][j - 1] + cost)  # substitution

    last_element = s.split("/")[-1]
    last_element_index = ''.join(filter(str.isdigit, last_element))
    print(last_element_index)
    last_element1 = t.split("/")[-1]
    last_element_index1 = ''.join(filter(str.isdigit, last_element1))
    print(last_element_index1)
    return d[m][n]

def verify_isLeaf(attribute,input,output):
    workbook = openpyxl.load_workbook(input)  # attr
    worksheet = workbook.active

    workbook_output = openpyxl.load_workbook(output)  # similarity
    worksheet_output = workbook_output.active
    for i in range(1, worksheet.max_row + 1):
        value0 = worksheet.cell(column=1, row=i).value

        if value0 == None:
            continue
        if '-' in value0:
            continue
        if value0[0].isupper():
            if attribute == 11:
                attribute += 2  # is leaf in column 15
            fileName = worksheet.cell(column=1, row=i).value
            isLeaf_target = worksheet.cell(column=attribute + 1, row=i).value  # isLeaf
            # isLeaf_distance = 1000000
            isLeaf_id = ''
            for j in range(1, 11):
                id_c = worksheet.cell(column=2, row=i + j).value
                if id_c == None:
                    print("candidates less than 10, id==None")
                    break
                isLeaf_c = worksheet.cell(column=attribute + 1, row=i + j).value
                if isLeaf_c == None:
                    isLeaf_c = ''
                # isLeaf_distance0 = levenshtein_distance(isLeaf_target, isLeaf_c)
                if isLeaf_c==isLeaf_target:
                    # isLeaf_distance = isLeaf_distance0
                    isLeaf_id += str(id_c).replace(".0", "") + ","
                # elif isLeaf_distance0 == isLeaf_distance:
                #     isLeaf_id += str(id_c).replace(".0", "") + ","
            for z in range(2, worksheet_output.max_row + 1):
                value1 = worksheet_output.cell(column=1, row=z).value
                if fileName == None or value1 == None:
                    continue
                if fileName in value1:
                    if attribute == 13:
                        attribute -= 2
                    worksheet_output.cell(column=attribute, row=z, value=isLeaf_id)
                    print("verify " + str(z))
                    workbook_output.save(output)
def verify_fill_empty(attribute,input,output):#location and xpath
    # output = "D:/sustech/0graduation/final/before_combine_20/similarities_verified.xlsx"
    # input = "D:/sustech/0graduation/final/before_combine_20/attributes.xlsx"
    workbook = openpyxl.load_workbook(input) #attr
    worksheet = workbook.active

    workbook_output=openpyxl.load_workbook(output) #similarity
    worksheet_output=workbook_output.active
    for i in range(1,worksheet.max_row+1):
        value0=worksheet.cell(column=1,row=i).value

        if value0==None:
            continue
        if '-' in value0:
            continue
        if value0[0].isupper():
            if attribute==11:
                attribute+=2#is leaf in column 15
            fileName=worksheet.cell(column=1,row=i).value
            attr_target=worksheet.cell(column=attribute+1,row=i).value#name
            attr_distance=1000000
            attr_id=''
            if "TestaddPhone" in fileName:
                continue
            for j in range(1,11):
                id_c = worksheet.cell(column=2,row=i+j).value
                if id_c==None:
                    print("candidates less than 10, id==None")
                    break
                attr_c = worksheet.cell(column=attribute+1, row=i+j).value
                if attr_c==None:
                    attr_c=''
                attr_distance0=levenshtein_distance(attr_target,attr_c)
                if attr_distance0<attr_distance:
                    attr_distance=attr_distance0
                    attr_id=str(id_c).replace(".0","")+","
                elif attr_distance0==attr_distance:
                    attr_id+=str(id_c).replace(".0","")+","
            for z in range(2,worksheet_output.max_row+1):
                value1=worksheet_output.cell(column=1,row=z).value
                if fileName==None or value1==None:
                    continue
                if fileName in value1:
                    if attribute==13:
                        attribute-=2
                    worksheet_output.cell(column=attribute,row=z,value=attr_id)
                    print("verify "+ str(z))
                    workbook_output.save(output)

def locationSim(attribute,input,output):
    workbook = openpyxl.load_workbook(input)
    worksheet = workbook.active
    workbook_output = openpyxl.Workbook()
    # workbook_output = openpyxl.load_workbook(output)
    worksheet_output = workbook_output.active
    worksheet_output.cell(row=1,column=1,value="fileName")
    worksheet_output.cell(row=1, column=2, value="id")
    worksheet_output.cell(row=1, column=3, value="name")
    worksheet_output.cell(row=1, column=4, value="class")
    worksheet_output.cell(row=1, column=5, value="xpath")
    worksheet_output.cell(row=1, column=6, value="text")
    worksheet_output.cell(row=1, column=7, value="tagName")
    # worksheet_output.cell(row=1, column=8, value="href")
    worksheet_output.cell(row=1, column=8, value="linkText")
    worksheet_output.cell(row=1, column=9, value="location")
    worksheet_output.cell(row=1, column=10, value="size")
    worksheet_output.cell(row=1, column=11, value="is leaf")
    worksheet_output.cell(row=1, column=12, value="gt")
    row_num=2
    for i in range(1, worksheet.max_row+1):
        value0 = worksheet.cell(column=1, row=i).value

        if value0 == None:
            continue
        if '-' in value0:
            continue
        if value0[0].isupper():
            print("att filename: " + value0)
            fileName = worksheet.cell(column=1, row=i).value
            x_target = worksheet.cell(column=attribute + 1, row=i).value  # x
            y_target = worksheet.cell(column=attribute + 2, row=i).value  # y
            h_target = worksheet.cell(column=attribute + 3, row=i).value  # height
            w_target = worksheet.cell(column=attribute + 4, row=i).value  # weight
            print(fileName+" line(attr): "+str(i))
            print(x_target+","+y_target+","+h_target+","+w_target)
            size_target = int(h_target)*int(w_target)
            worksheet_output.cell(column=1,row=row_num,value=fileName)

            min_distance = 1000000#location
            min_size = 10000000000
            location_id = ''#location
            size_id = ''
            print(input)
            print(output)
            for j in range(1, 11):
                filename_c = worksheet.cell(column=1, row=i+j).value
                if filename_c!=None:
                    print("candidates less than 10, filename!=None")
                    break
                id_c = worksheet.cell(column=2, row=i + j).value
                x_c = worksheet.cell(column=attribute + 1, row=i + j).value
                y_c = worksheet.cell(column=attribute + 2, row=i + j).value
                h_c = worksheet.cell(column=attribute + 3, row=i + j).value
                w_c = worksheet.cell(column=attribute + 4, row=i + j).value
                # if name_c == None:
                #     name_c = ''
                if id_c==None:
                    print("candidates less than 10, id==None")
                    break
                print("line "+str(i+j))
                cur_distance=points_distance(x_c,y_c,x_target,y_target)
                cur_size=int(h_c)*int(w_c)
                print(fileName+" "+id_c)
                if math.fabs(cur_size-size_target)<min_size:
                    min_size=math.fabs(cur_size-size_target)
                    size_id=str(id_c).replace(".0", "") + ","
                if cur_distance < min_distance:
                    min_distance = cur_distance
                    location_id= str(id_c).replace(".0", "") + ","
                elif cur_distance == min_distance:
                    location_id+= str(id_c).replace(".0", "") + ","
            worksheet_output.cell(column=attribute, row=row_num, value=location_id)
            worksheet_output.cell(column=attribute+1, row=row_num, value=size_id)
            row_num+=1
            print()
        row_num+=1
    workbook_output.save(output)

def getGTID(attributes,similarities,gt_path):
    workbook_attr = openpyxl.load_workbook(attributes)
    worksheet_attr = workbook_attr.active
    workbook_sim = openpyxl.load_workbook(similarities)
    worksheet_sim = workbook_sim.active

    for i in range(1,worksheet_sim.max_row+1):
        filename0=worksheet_sim.cell(row=i,column=1).value

        if filename0==None or "_" not in filename0:
            continue
        filename0 = filename0.replace("_all.txt", "")
        # print("filename0: "+filename0)
        for root, dirs, files in os.walk(gt_path):
            for filename in files:
                # print("filename: "+filename)
                if filename.endswith(".txt") and filename0 in filename:
                    filepath = os.path.join(root, filename)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = f.read()
                    datas = data.split("\n")
                    for content in datas:
                        if "expect xpath:" in content:
                            gt_xpath = content.split(":")[1]
                            print("expect xpath:"+gt_xpath)
                            break
                        else:
                            gt_xpath=datas[0]
        for j in range(1,worksheet_attr.max_row+1):
            filename_att=worksheet_attr.cell(row=j,column=1).value
            if filename_att == None or "_" not in filename_att:
                continue
            # filename=filename.replace(".txt","")
            if filename0 in filename_att:
                for z in range(1,11):
                    c_xpath=worksheet_attr.cell(row=j+z,column=6).value
                    c_id=worksheet_attr.cell(row=j+z,column=2).value
                    if c_id==None:
                        continue
                    c_xpath=c_xpath.replace("'","").replace("[1]","")
                    gt_xpath=gt_xpath.replace("[1]","")
                    print("c_xpath: "+c_xpath)
                    if c_xpath==gt_xpath:
                        print("find gt")
                        worksheet_sim.cell(row=i,column=13,value=c_id)
                        break
                    else:
                        print(gt_path)
                        print(c_xpath)
                        print()
                break
    workbook_sim.save(similarities)


def getGTInCandidate(attributes,similarities,gt_path,re):
    workbook_re = openpyxl.load_workbook(re)
    worksheet_re = workbook_re.active
    workbook_attr = openpyxl.load_workbook(attributes)
    worksheet_attr = workbook_attr.active
    workbook_sim = openpyxl.load_workbook(similarities)
    worksheet_sim = workbook_sim.active

    for i in range(1,worksheet_re.max_row+2):
        filename0=worksheet_re.cell(row=i,column=1).value

        if filename0==None or "_" not in filename0:
            continue
        filename0 = filename0.replace("_all.txt", "")
        # print("filename0: "+filename0)
        for root, dirs, files in os.walk(gt_path):
            for filename in files:
                # print("filename: "+filename)
                if filename.endswith(".txt") and filename0 in filename:
                    filepath = os.path.join(root, filename)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = f.read()
                    datas = data.split("\n")
                    for content in datas:
                        if "expect xpath:" in content:
                            gt_xpath = content.split(":")[1]
                            print("expect xpath:"+gt_xpath)
                            break
                        else:
                            gt_xpath=datas[0].replace("[1]","")
        for j in range(1,worksheet_attr.max_row+1):
            filename_att=worksheet_attr.cell(row=j,column=1).value
            if filename_att == None or "_" not in filename_att:
                continue
            # filename=filename.replace(".txt","")
            if filename0 in filename_att:
                for z in range(1,11):
                    c_xpath=worksheet_attr.cell(row=j+z,column=5).value
                    c_id=worksheet_attr.cell(row=j+z,column=2).value
                    if c_id==None:
                        continue
                    c_xpath=c_xpath.replace("'","").replace("[1]","")
                    print("c_xpath: "+c_xpath)
                    if c_xpath==gt_xpath:
                        print("find gt")
                        worksheet_re.cell(row=i,column=2,value=1)
                        worksheet_re.cell(row=i, column=3, value=z)
                        break
                    else:
                        print(gt_path)
                        print(c_xpath)
                        print()
                break
    workbook_re.save(re)


def generateSimilarities(input,output):
    locationSim(9, input, output)  # location and size
    verify_fill_empty(2, input, output)  # id
    verify_fill_empty(3, input, output)  # name
    verify_fill_empty(4, input, output)  # class
    verify_fill_empty(5, input, output)  # xpath
    verify_fill_empty(6, input, output)  # text
    verify_fill_empty(7, input, output)  # tagName
    # verify_fill_empty(8, input, output)  # href
    verify_fill_empty(8, input, output)  # link text
    # verify_fill_empty(8, input, output)  # location 8,9
    # verify_fill_empty(9, input, output)  # size 10,11
    verify_isLeaf(11, input, output)  # is leaf
if __name__ == "__main__":
    #output:similarity input:attribute
    gt_paths = "\\ground_truth\\"
    prefix = ""
    methods = ["vista", "water", "xpath"]
    for method in methods:
        attr = method + "_attributes.xlsx"
        sim = method + "_similarities.xlsx"
        output = method + "_match.xlsx"
        generateSimilarities(attr, sim)
        getGTID(attr, output, gt_paths)


