import os

import openpyxl

def generate_attributes(path,output):
    workbook = openpyxl.Workbook()
    # workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    row_num = 2
    sheet.cell(row=1, column=1, value="file name")
    sheet.cell(row=1, column=2, value="numeric id")
    sheet.cell(row=1, column=3, value="id")
    sheet.cell(row=1, column=4, value="name")
    sheet.cell(row=1, column=5, value="class")
    sheet.cell(row=1, column=6, value="xpath")
    sheet.cell(row=1, column=7, value="text")
    sheet.cell(row=1, column=8, value="tagName")
    sheet.cell(row=1, column=9, value="link text")
    sheet.cell(row=1, column=10, value="x")
    sheet.cell(row=1, column=11, value="y")
    sheet.cell(row=1, column=12, value="width")
    sheet.cell(row=1, column=13, value="height")
    sheet.cell(row=1, column=14, value="is leaf")
    sheet.cell(row=1, column=15, value="is target")
    row_num=2
    for root, dirs, files in os.walk(path):
        for filename in files:
            if ".txt" in filename:
                filepath = os.path.join(root, filename)
                print(filename)
                sheet.cell(row=row_num, column=1, value=filename)
                sheet.cell(row=row_num, column=15, value="is target")
                with open(filepath, 'r',encoding='utf-8') as f:
                    data = f.read()
                target, candidates=getElementInfo(data)
                t_c=target+candidates
                print(t_c)
                elements = t_c.split("}")
                for e in elements:
                    print(len(elements))
                    attrs=e.split(",")
                    col=2
                    for attr in attrs:
                        if "=" in attr:
                            print(attr)
                            values=attr.split("=")
                            sheet.cell(row=row_num,column=col,value=values[1])
                            col+=1
                    row_num+=1
                row_num+=2
    workbook.save(output)



def getElementInfo(data):
    prefix1="Target element: "
    index1=data.find(prefix1)+len(prefix1)
    # print(index1)
    prefix2="Candidate elements: "
    index2=data.find(prefix2)
    # print(index2)
    target=data[index1:index2]
    prefix3="Time: "
    index3 = data.find(prefix3)
    # print(index3)
    candidates=data[index2+len(prefix2):index3]
    return target,candidates

if __name__ == "__main__":
    method=""
    candidates_path="\\extracted_element\\"+method
    output_path=method+"_attributes.xlsx"
    generate_attributes(candidates_path,output_path)