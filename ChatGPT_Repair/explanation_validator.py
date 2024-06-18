import os
import time
import openai
import json
import openpyxl

def process_files1(path,output):
    a1="attribute"
    a2="structure"
    a3="without"

    for root, dirs, files in os.walk(path):
        for filename in files:
            if a1 in filename:
                filepath = os.path.join(root, filename)
                with open(filepath, 'r') as f:
                    data = f.read()
                    # print(data)
                #extract answer id
                index=[100000]*10
                index[0]=100000
                index[1]=find_index(data,",")
                index[2]=find_index(data,".")
                index[3]=find_index(data," because")
                index[4]=find_index(data," for")
                index[5]=find_index(data," since")
                index[6]=find_index(data," based on")
                index[7]=find_index(data," according to")
                index[8]=find_index(data," as")
                index[9]=find_index(data," which")

                if "Based on" in data:
                    sub_str=data[index[1]:index[2]]
                else:
                    min_index = min(index)
                    sub_str = data[:min_index]
                # print(sub_str) # find the substring contains answer id
                index_is=find_index(sub_str,"is ")
                # print(index_is)
                if index_is!=10000:
                    sub_str=sub_str[index_is:]
                    #print(sub_str)
                sub_str=sub_str.replace("{","")
                sub_str=sub_str.replace("}","")
                index2=[10000]*9
                index2[0]=find_index2(sub_str,"elementId=")
                index2[1]=find_index2(sub_str,"element ")
                index2[2]=find_index2(sub_str,"ElementId: ")
                index2[3]=find_index2(sub_str,"id=")
                min_index=min(index2)
                # print(min_index)
                sub_str2=sub_str[min_index:]
                print(sub_str2)
                print("------------")
                sub_index=sub_str.find("elementId=",sub_str.find("elementId=")+len("elementId="))
                if sub_index!=-1:#find the answer id
                    answerId=sub_str[sub_index:]
                else:
                    sub_index = sub_str.find("is ", sub_str.find("is ") + len("is "))
                    if sub_index != -1:  # find the answer id
                        answerId = sub_str[sub_index:]
                    else:
                        # print("didn't find the answer's id\r\n"+data+"----------------")
                        # print("not found")
                        continue
                # print(answerId)
                #extract attributes considered

                # check whether un-wanted attributes

                #calculate right and unknown attributes



                #check whether it is ground-truth

                #store in excel
                # # Open the workbook and select the sheet
                # workbook = openpyxl.load_workbook(output)
                # sheet = workbook.active
                #
                # # Check if a row is empty
                # for row in sheet.iter_rows(min_row=1, max_row=10):
                #     if all(cell.value is None for cell in row):
                #         print(f"Row {row[0].row} is empty")
                #
                # # Check if a column is empty
                # for column in sheet.iter_cols(min_col=1, max_col=10):
                #     if all(cell.value is None for cell in column):
                #         print(f"Column {column[0].column_letter} is empty")
                #
                # # Write data to a specific cell
                # sheet['A1'] = 'Hello World!'
                #
                # # Save the changes to the workbook
                # workbook.save('example.xlsx')

def check_attributes(path,output):
    a1 = "attribute"
    a2 = "structure"
    a3 = "without"
    # prefix_output="D:/sustech/0graduation/final/before_combine_20/chatgpt_result/analysis_"
    # output1=prefix_output+a1+".xlsx"
    # output2 = prefix_output + a2 + ".xlsx"
    # output3 = prefix_output + a3 + ".xlsx"
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    row_num = 2

    for root, dirs, files in os.walk(path):
        for filename in files:
            if a1 in filename and filename.endswith(".txt"):
                filepath = os.path.join(root, filename)
                str_arr = filename.split("_")
                name = str_arr[0] + "_" + str_arr[1]
                # print(filepath)
                with open(filepath, 'r') as f:
                    data = f.read()
                    # print(data)
                # extract answer id
                index = [100000] * 10
                index[0] = 100000
                index[1] = find_index(data, ",")
                index[2] = find_index(data, ".")
                index[3] = find_index(data, " because")
                index[4] = find_index(data, " for")
                index[5] = find_index(data, " since")
                index[6] = find_index(data, " based on")
                index[7] = find_index(data, " according to")
                index[8] = find_index(data, " as")
                index[9] = find_index(data, " which")

                if "Based on" in data:
                    sub_str = data[:index[1]]
                else:
                    min_index = min(index)
                    sub_str = data[min_index:]

                sub_str = sub_str.replace("{", "")
                sub_str = sub_str.replace("}", "")
                index2 = [10000] * 9
                index2[0] = find_index3(sub_str, "repair")
                index2[1] = find_index3(sub_str, "Repair")
                index2[2] =find_index(sub_str,"fix")
                # index2[2] = find_index3(sub_str,".")
                min_index = min(index2)
                # print(min_index)
                sub_str2 = sub_str[:min_index]

                # print(sub_str2) # the sentence contains attributes

                # print("-------------")

                id=sheet.cell(row=row_num,column=2).value
                print("line "+str(row_num)+" in match, filename:"+filename)
                # print(str(id))
                # for i in range(1,sheet.max_row):
                #     cell_value=sheet.cell(row=i,column=1)
                #     if name in cell_value.value:
                #         id=sheet.cell(row=i,column=2).value
                #         this_row=i
                #         break

                # for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
                #     if name in row:
                #         print(row[0].row)
                #         this_row=row[0]
                #
                #         id=sheet.cell(row=this_row.row, column=1)
                #         break
                # print("chatgpt result: "+id)
                attributes=[0]*10
                results=[0]*10

                total_num=0
                invalid_num=0
                correct_num=0
                #id, name, text, tagName, linkText, site, xpath, location, ground_truth
                if " id " in sub_str2:
                    results[0]=check_correct(name,id,1,row_num)
                    attributes[0]=1
                    total_num+=1
                if " name " in sub_str2:
                    results[1] = check_correct(name, id, 2,row_num)
                    attributes[1]=1
                    total_num += 1


                if " text " in sub_str2:
                    results[2] = check_correct(name, id, 3,row_num)
                    attributes[2]=1
                    total_num += 1
                if "tagName" in sub_str2 or "tag " in sub_str2:
                    results[3] = check_correct(name, id, 4, row_num)
                    attributes[3]=1
                    total_num+=1
                if "linkText" in sub_str2 or "link text" in sub_str2 or "link" in sub_str2:
                    results[4] = check_correct(name, id, 5,row_num)
                    attributes[4]=1
                    total_num += 1
                if "width" in sub_str2 or "height" in sub_str2 or "size" in sub_str2 or "dimension" in sub_str2:
                    results[5] = check_correct(name, id, 6,row_num)
                    attributes[5]=1
                    total_num += 1
                if "xpath" in sub_str2:
                    results[6] = check_correct(name, id, 7, row_num)
                    invalid_num+=1
                    attributes[6]=1
                    total_num += 1
                if " x " in sub_str2 or " y " in sub_str2 or "location" in sub_str2 or "layout" in sub_str2:
                    results[7] = check_correct(name, id, 8, row_num)
                    invalid_num+=1
                    attributes[7]=1
                    total_num += 1
                if "same " in sub_str2:
                    same=1
                else:
                    same=0
                if "similar " in sub_str2:
                    similar=1
                else:
                    similar=0
                this_row = row_num
                sheet.cell(row=this_row,column=27,value=same)
                sheet.cell(row=this_row,column=28,value=similar)

                correct_num=sum(results)

                for i in range(3,11):
                    # this_cell[i]=attributes[i-2]
                    sheet.cell(row=this_row, column=i, value=attributes[i-3])
                for i in range(13,21):
                    # this_cell[i]=results[i-12]
                    sheet.cell(row=this_row, column=i, value=results[i - 13])
                # cell[22]=total_num
                # cell[23]=invalid_num
                # cell[24]=total_num-invalid_num
                # cell[25]=correct_num


                # sheet.cell(row=row_num,column=1,value=name)
                # # row_num+=1
                # sheet.cell(row=row_num,column=2,value=re)
                # row_num+=1

                # # Save the changes to the workbook
                workbook.save(output)
                row_num+=1


def check_structure(path,output):
    a1 = "attribute"
    a2 = "structure"
    a3 = "without"
    # prefix_output="D:/sustech/0graduation/final/before_combine_20/chatgpt_result/analysis_"
    # output1=prefix_output+a1+".xlsx"
    # output2 = prefix_output + a2 + ".xlsx"
    # output3 = prefix_output + a3 + ".xlsx"
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    row_num = 2

    for root, dirs, files in os.walk(path):
        for filename in files:
            if a2 in filename and filename.endswith(".txt"):
                filepath = os.path.join(root, filename)
                str_arr = filename.split("_")
                name = str_arr[0] + "_" + str_arr[1]
                # print(filepath)
                with open(filepath, 'r') as f:
                    data = f.read()
                    # print(data)
                # extract answer id
                index = [100000] * 10
                index[0] = 100000
                index[1] = find_index(data, ",")
                index[2] = find_index(data, ".")
                index[3] = find_index(data, " because")
                index[4] = find_index(data, " for")
                index[5] = find_index(data, " since")
                index[6] = find_index(data, " based on")
                index[7] = find_index(data, " according to")
                index[8] = find_index(data, " as")
                index[9] = find_index(data, " which")

                if "Based on" in data:
                    sub_str = data[:index[1]]
                else:
                    min_index = min(index)
                    sub_str = data[min_index:]

                sub_str = sub_str.replace("{", "")
                sub_str = sub_str.replace("}", "")
                index2 = [10000] * 9
                index2[0] = find_index3(sub_str, "repair")
                index2[1] = find_index3(sub_str, "Repair")
                index2[2] =find_index(sub_str,"fix")
                # index2[2] = find_index3(sub_str,".")
                min_index = min(index2)
                # print(min_index)
                sub_str2 = sub_str[:min_index]

                # print(sub_str2) # the sentence contains attributes

                # print("-------------")

                id=sheet.cell(row=row_num,column=2).value

                attributes=[0]*10
                results=[0]*10

                total_num=0
                invalid_num=0
                correct_num=0
                #id, name, text, tagName, linkText, site, xpath, location, ground_truth
                if " id " in sub_str2:
                    results[0]=check_correct(name,id,1,row_num)
                    attributes[0]=1
                    total_num+=1
                if " name " in sub_str2:
                    results[1] = check_correct(name, id, 2,row_num)
                    attributes[1]=1
                    total_num += 1


                if " text " in sub_str2:
                    results[2] = check_correct(name, id, 3,row_num)
                    attributes[2]=1
                    total_num += 1
                if "tagName" in sub_str2 or "tag " in sub_str2:
                    results[3] = check_correct(name, id, 4, row_num)
                    attributes[3]=1
                    total_num+=1
                if "linkText" in sub_str2 or "link text" in sub_str2 or "link" in sub_str2:
                    results[4] = check_correct(name, id, 5,row_num)
                    attributes[4]=1
                    total_num += 1
                if "width" in sub_str2 or "height" in sub_str2 or "size" in sub_str2 or "dimension" in sub_str2:
                    results[5] = check_correct(name, id, 6,row_num)
                    attributes[5]=1
                    total_num += 1
                if "xpath" in sub_str2 or "parent" in sub_str2 or "children" in sub_str2:
                    results[6] = check_correct(name, id, 7, row_num)
                    # invalid_num+=1
                    attributes[6]=1
                    total_num += 1
                if " x " in sub_str2 or " y " in sub_str2 or "location" in sub_str2 or "layout" in sub_str2:
                    results[7] = check_correct(name, id, 8, row_num)
                    # invalid_num+=1
                    attributes[7]=1
                    total_num += 1
                if "same " in sub_str2:
                    same=1
                else:
                    same=0
                if "similar " in sub_str2:
                    similar=1
                else:
                    similar=0
                this_row = row_num
                sheet.cell(row=this_row,column=27,value=same)
                sheet.cell(row=this_row,column=28,value=similar)
                correct_num=sum(results)
                # this_row=row_num
                for i in range(3,11):
                    # this_cell[i]=attributes[i-2]
                    sheet.cell(row=this_row, column=i, value=attributes[i-3])
                for i in range(13,21):
                    # this_cell[i]=results[i-12]
                    sheet.cell(row=this_row, column=i, value=results[i - 13])
                # cell[22]=total_num
                # cell[23]=invalid_num
                # cell[24]=total_num-invalid_num
                # cell[25]=correct_num


                # sheet.cell(row=row_num,column=1,value=name)
                # # row_num+=1
                # sheet.cell(row=row_num,column=2,value=re)
                # row_num+=1

                # # Save the changes to the workbook
                workbook.save(output)
                row_num+=1

def check_without(path,output):
    a1 = "attribute"
    a2 = "structure"
    a1 = "without"
    # prefix_output="D:/sustech/0graduation/final/before_combine_20/chatgpt_result/analysis_"
    # output1=prefix_output+a1+".xlsx"
    # output2 = prefix_output + a2 + ".xlsx"
    # output3 = prefix_output + a3 + ".xlsx"
    workbook = openpyxl.load_workbook(output)
    # print("open "+output)
    sheet = workbook.active
    row_num = -1
    print(path)
    print(output)
    for root, dirs, files in os.walk(path):
        for filename in files:
            # print(filename)
            # print(a)
            if filename.endswith(".txt"):
                filepath = os.path.join(root, filename)
                str_arr = filename.split("_")
                name = str_arr[0] + "_" + str_arr[1]
                # print(filepath)
                for i in range(1,sheet.max_row+2):
                    output_filename=sheet.cell(row=i,column=1).value
                    if output_filename==filename.replace("_answer.txt",""):
                        row_num=i
                if row_num==-1:
                    print("didn't find the row of this txt "+filename)
                    continue
                with open(filepath, 'r', encoding="utf-8") as f:
                    data = f.read()
                    # print(data)
                # extract answer id
                try:
                    index1=data.index("attributes")
                except ValueError:
                    print(filepath)
                    print("cant find attributes")
                    continue
                try:
                    index2=data.index("driver")
                    sub_str=data[:index1]
                    sub_str2=data[index1:index2]
                    print(sub_str)
                    print(sub_str2)
                    print("---")
                except ValueError:
                    sub_str = data[:index1]
                    sub_str2 = data[index1:]

                sub_str = sub_str.replace("{", "")
                sub_str = sub_str.replace("}", "")
                index2 = [10000] * 9
                index2[0] = find_index3(sub_str, "repair")
                index2[1] = find_index3(sub_str, "Repair")
                index2[2] =find_index(sub_str,"fix")
                # index2[2] = find_index3(sub_str,".")
                min_index = min(index2)
                # print(min_index)


                # print(sub_str2) # the sentence contains attributes

                # print("-------------")
                # row_num=
                id=sheet.cell(row=row_num,column=2).value
                print("line: "+str(row_num)+" in match file, filename: "+filename)
                attributes=[0]*11
                results=[0]*11

                total_num=0
                invalid_num=0
                correct_num=0
                #id, name, text, tagName, linkText, site, xpath, location, ground_truth
                if " id " in sub_str2 or " id." in sub_str2 or " id," in sub_str2:
                    results[0]=check_correct(name,id,1,row_num,method)
                    attributes[0]=1
                    total_num+=1
                if " name " in sub_str2 or " name." in sub_str2 or " name," in sub_str2:
                    results[1] = check_correct(name, id, 2,row_num,method)
                    attributes[1]=1
                    total_num += 1
                if " class " in sub_str2 or " class." in sub_str2 or " class," in sub_str2:
                    results[2] = check_correct(name, id, 3,row_num,method)
                    attributes[2]=1
                    total_num += 1
                if "xpath" in sub_str2:
                    results[3] = check_correct(name, id, 4, row_num, method)
                    # invalid_num+=1
                    attributes[3] = 1
                    total_num += 1

                if " text " in sub_str2 or " text." in sub_str2 or "text," in sub_str2:
                    results[4] = check_correct(name, id, 5,row_num,method)
                    attributes[4]=1
                    total_num += 1
                if "tagName" in sub_str2 or "tag " in sub_str2 or "tag." in sub_str2 or " tag," in sub_str2 or "tagName." in sub_str2 or " tagName," in sub_str2:
                    results[5] = check_correct(name, id, 6, row_num,method)
                    attributes[5]=1
                    total_num+=1
                # if "href" in sub_str2:
                #     results[6] = check_correct(name, id, 7,row_num,method)
                #     attributes[6]=1
                #     total_num += 1
                if "linkText" in sub_str2 or "link text" in sub_str2 or "link" in sub_str2:
                    results[6] = check_correct(name, id, 7,row_num,method)
                    attributes[6]=1
                    total_num += 1
                if "width" in sub_str2 or "height" in sub_str2 or "size" in sub_str2 or "dimension" in sub_str2:
                    results[7] = check_correct(name, id, 8,row_num,method)
                    attributes[7]=1
                    total_num += 1

                if " x " in sub_str2 or " y " in sub_str2 or "location" in sub_str2 or "layout" in sub_str2 or " x." in sub_str2 or " y." in sub_str2 or " x," in sub_str2 or " y," in sub_str2:
                    results[8] = check_correct(name, id, 9, row_num,method)
                    # invalid_num+=1
                    attributes[8]=1
                    total_num += 1
                if "isLeaf" in sub_str2 or "is leaf" in sub_str2 or "leaf" in sub_str2:
                    results[9] = check_correct(name,id,10,row_num,method)
                    attributes[9] = 1
                    total_num += 1
                if "same " in sub_str2:
                    same=1
                else:
                    same=0
                if "similar " in sub_str2:
                    similar=1
                else:
                    similar=0
                this_row = row_num
                # sheet.cell(row=this_row,column=27,value=same)
                # sheet.cell(row=this_row,column=28,value=similar)
                correct_num=sum(results)
                sheet.cell(row=this_row,column=27,value=total_num)
                sheet.cell(row=this_row,column=28,value=correct_num)
                if total_num!=0:
                    sheet.cell(row=this_row,column=29,value=correct_num/total_num)
                else:
                    sheet.cell(row=this_row, column=29, value=0)
                # this_row=row_num
                for i in range(3,13):
                    # this_cell[i]=attributes[i-2]
                    sheet.cell(row=this_row, column=i, value=attributes[i-3])
                for i in range(15,25):
                    # this_cell[i]=results[i-12]
                    sheet.cell(row=this_row, column=i, value=results[i - 15])
                # cell[22]=total_num
                # cell[23]=invalid_num
                # cell[24]=total_num-invalid_num
                # cell[25]=correct_num


                # sheet.cell(row=row_num,column=1,value=name)
                # # row_num+=1
                # sheet.cell(row=row_num,column=2,value=re)
                # row_num+=1

                # # Save the changes to the workbook
                workbook.save(output)
                # print("save "+output)
                row_num+=1
def check_without0(path,output,):
    a1 = "attribute"
    a2 = "structure"
    a1 = "without"
    # prefix_output="D:/sustech/0graduation/final/before_combine_20/chatgpt_result/analysis_"
    # output1=prefix_output+a1+".xlsx"
    # output2 = prefix_output + a2 + ".xlsx"
    # output3 = prefix_output + a3 + ".xlsx"
    workbook = openpyxl.load_workbook(output)
    # print("open "+output)
    sheet = workbook.active
    row_num = 2

    for root, dirs, files in os.walk(path):
        for filename in files:
            # print(filename)
            # print(a)
            if filename.endswith(".txt") :
                filepath = os.path.join(root, filename)
                str_arr = filename.split("_")
                name = str_arr[0] + "_" + str_arr[1]
                # print(filepath)
                with open(filepath, 'r') as f:
                    data = f.read()
                    # print(data)
                # extract answer id
                index = [100000] * 10
                index[0] = 100000
                index[1] = find_index(data, " Because")
                index[2] = find_index(data, ".")
                index[3] = find_index(data, " because")
                index[4] = find_index(data, " for")
                index[5] = find_index(data, " since")
                index[6] = find_index(data, " based on")
                index[7] = find_index(data, " according to")
                index[8] = find_index(data, " as")
                index[9] = find_index(data, " which")

                if "Based on" in data:
                    sub_str = data[:index[1]]
                else:
                    min_index = min(index)
                    sub_str = data[min_index:]

                sub_str = sub_str.replace("{", "")
                sub_str = sub_str.replace("}", "")
                index2 = [10000] * 9
                index2[0] = find_index3(sub_str, "repair")
                index2[1] = find_index3(sub_str, "Repair")
                index2[2] =find_index(sub_str,"fix")
                # index2[2] = find_index3(sub_str,".")
                min_index = min(index2)
                # print(min_index)
                sub_str2 = sub_str[:min_index]

                # print(sub_str2) # the sentence contains attributes

                # print("-------------")

                id=sheet.cell(row=row_num,column=2).value

                attributes=[0]*10
                results=[0]*10

                total_num=0
                invalid_num=0
                correct_num=0
                #id, name, text, tagName, linkText, site, xpath, location, ground_truth
                if " id " in sub_str2:
                    results[0]=check_correct(name,id,1,row_num,method)
                    attributes[0]=1
                    total_num+=1
                if " name " in sub_str2:
                    results[1] = check_correct(name, id, 2,row_num,method)
                    attributes[1]=1
                    total_num += 1


                if " text " in sub_str2:
                    results[2] = check_correct(name, id, 3,row_num,method)
                    attributes[2]=1
                    total_num += 1
                if "tagName" in sub_str2 or "tag " in sub_str2:
                    results[3] = check_correct(name, id, 4, row_num,method)
                    attributes[3]=1
                    total_num+=1
                if "linkText" in sub_str2 or "link text" in sub_str2 or "link" in sub_str2:
                    results[4] = check_correct(name, id, 5,row_num,method)
                    attributes[4]=1
                    total_num += 1
                if "width" in sub_str2 or "height" in sub_str2 or "size" in sub_str2 or "dimension" in sub_str2:
                    results[5] = check_correct(name, id, 6,row_num,method)
                    attributes[5]=1
                    total_num += 1
                if "xpath" in sub_str2 or "parent" in sub_str2 or "children" in sub_str2:
                    results[6] = check_correct(name, id, 7, row_num,method)
                    # invalid_num+=1
                    attributes[6]=1
                    total_num += 1
                if " x " in sub_str2 or " y " in sub_str2 or "location" in sub_str2 or "layout" in sub_str2:
                    results[7] = check_correct(name, id, 8, row_num,method)
                    # invalid_num+=1
                    attributes[7]=1
                    total_num += 1
                if "same " in sub_str2:
                    same=1
                else:
                    same=0
                if "similar " in sub_str2:
                    similar=1
                else:
                    similar=0
                this_row = row_num
                sheet.cell(row=this_row,column=27,value=same)
                sheet.cell(row=this_row,column=28,value=similar)
                correct_num=sum(results)
                # this_row=row_num
                for i in range(3,11):
                    # this_cell[i]=attributes[i-2]
                    sheet.cell(row=this_row, column=i, value=attributes[i-3])
                for i in range(13,21):
                    # this_cell[i]=results[i-12]
                    sheet.cell(row=this_row, column=i, value=results[i - 13])
                # cell[22]=total_num
                # cell[23]=invalid_num
                # cell[24]=total_num-invalid_num
                # cell[25]=correct_num


                # sheet.cell(row=row_num,column=1,value=name)
                # # row_num+=1
                # sheet.cell(row=row_num,column=2,value=re)
                # row_num+=1

                # # Save the changes to the workbook
                workbook.save(output)
                # print("save "+output)
                row_num+=1


def check_correct(filename,id,attribute,row_num,method):
    similarity_path="D:\\Concordia\\lab search\\graduation_papper_undergraduate\\without_history1109_4_times\\"+method+"_similarities.xlsx"
    # similarity_path = "D:/sustech/0graduation/final/before_add/similarities_verified.xlsx"
    workbook = openpyxl.load_workbook(similarity_path)
    # print("check correct open "+similarity_path)
    sheet = workbook.active
    # cell_value=sheet.cell(column=1,row=row_num).value
    # answers=sheet.cell(row=row_num,column=attribute+1).value
    answers=''
    for i in range(1,sheet.max_row+1):
        cell_value=sheet.cell(column=1,row=i).value
        if cell_value==None:
            continue
        if filename in cell_value:
            print(filename+" "+cell_value+" "+ str(attribute))
            answers=sheet.cell(row=i,column=attribute+1).value
            print("answers"+ str(answers)+" selected id: "+str(id))

            if answers=='' or answers==None:
                print("answer is none")
                return 0
            if type(answers)==int:
                print("answer is int type, value: "+str(answers))
                if id==answers:
                    return 1
                else:
                    return 0

            answersss=answers.split(",")

    # for cell in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
    #     if filename in cell[0]:
    #         answers=cell[attribute]
    #         break
    # print(answers)
            print("line 690 id: "+ str(id))
            print(answersss)
            for answer in answersss:
                if answer=='':
                    continue
                if answer==None:
                    continue
                if answer=="None":
                    continue
                if id==None:
                    continue
                if id=='null':
                    continue
                print(id)
                print(answer)
                if int(id)==int(answer) or str(id)==str(answer):
                    print("explain correctly")
                    return 1
                else:
                    print(str(id)+"!="+str(answer))
            print("explain wrongly")
            print("------")
            # if "," not in str(answers) and str(id) in str(answers):
            #     return 1
            # # if ","+str(id)+"," in answers or " "+str(id)+"," in answers :
            # if "," + str(id) + "," in answers or " " + str(id) + "," in answers:
            #     # print( str(id) +" is in answers: "+ answers)
            #     return 1

    return 0

def check_correct0(filename,id,attribute,row_num,similarity_path):
    # similarity_path="D:/sustech/0graduation/final/chatgpt10_0520/xp/similarities_xp.xlsx"
    # similarity_path = "D:/sustech/0graduation/final/before_add/similarities_verified.xlsx"
    workbook = openpyxl.load_workbook(similarity_path)
    # print("check correct open "+similarity_path)
    sheet = workbook.active
    # cell_value=sheet.cell(column=1,row=row_num).value
    # answers=sheet.cell(row=row_num,column=attribute+1).value
    answers=''
    for i in range(1,sheet.max_row+1):
        cell_value=sheet.cell(column=1,row=i).value
        if filename in cell_value:
            print(filename+" "+cell_value+" "+ str(attribute))
            answers=sheet.cell(row=i,column=attribute).value
            print("answers"+ str(answers)+" selected id: "+str(id))

            if answers=='' or answers==None:
                return 2
            if type(answers)==int:
                if id==answers:
                    return 1
                else:
                    return 0

            answersss=answers.split(",")

    # for cell in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
    #     if filename in cell[0]:
    #         answers=cell[attribute]
    #         break
    # print(answers)
            print(id)
            print("------")
            for answer in answersss:
                if id==answer:
                    return 1
            # if "," not in str(answers) and str(id) in str(answers):
            #     return 1
            # # if ","+str(id)+"," in answers or " "+str(id)+"," in answers :
            # if "," + str(id) + "," in answers or " " + str(id) + "," in answers:
            #     # print( str(id) +" is in answers: "+ answers)
            #     return 1

    return 0

def write_answer_id0(path,output,a):
    a1 = "attribute"
    a2 = "structure"
    a3 = "without"
    # prefix_output="D:/sustech/0graduation/final/before_combine_20/chatgpt_result/analysis_"
    # output1=prefix_output+a1+".xlsx"
    # output2 = prefix_output + a2 + ".xlsx"
    # output3 = prefix_output + a3 + ".xlsx"
    print(output)
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    sheet['A1'] = "file_name"
    sheet['B1'] = "selected_id"
    sheet['C1'] = "name"
    sheet['D1'] = "text"
    sheet['E1'] = "tagName"
    sheet['F1'] = "linkText"
    sheet['G1'] = "site"
    sheet['H1'] = "xpath"
    sheet['I1'] = "location"
    sheet['J1'] = "ground_truth"
    sheet['K1'] = ""
    sheet['L1'] = "mention"
    sheet['M1'] = "invalid"
    sheet['N1'] = "valid"
    sheet['O1'] = "correct"
    sheet['P1'] = "same"
    sheet['Q1'] = "similar"
    row_num=2
    for root, dirs, files in os.walk(path):
        for filename in files:
            if a in filename and filename.endswith(".txt"):
                filepath = os.path.join(root, filename)
                str_arr=filename.split("_")
                name=str_arr[0]+"_"+str_arr[1]
                # print(filepath)
                with open(filepath, 'r') as f:
                    data = f.read()
                    # print(data)
                # extract answer id
                index = [100000] * 10
                index[0] = 100000
                index[1] = find_index(data, ",")
                index[2] = find_index(data, ".")
                index[3] = find_index(data, " because")
                index[4] = find_index(data, " for")
                index[5] = find_index(data, " since")
                index[6] = find_index(data, " based on")
                index[7] = find_index(data, " according to")
                index[8] = find_index(data, " as")
                index[9] = find_index(data, " which")

                if "Based on" in data:
                    sub_str = data[index[1]:index[2]]
                else:
                    min_index = min(index)
                    sub_str = data[:min_index]
                # print(sub_str) # find the substring contains answer id
                index_is = find_index(sub_str, "is ")
                # print(index_is)
                if index_is != 10000:
                    sub_str = sub_str[index_is:]
                    print(sub_str)
                # sub_str = sub_str.replace("{", "")
                # sub_str = sub_str.replace("}", "")
                index2 = [10000] * 12
                index2[0] = find_index2(sub_str, "numericid=") + len("numericid=")
                # index2[1] = find_index2(sub_str, "element ") + len("element ")
                index2[2] = find_index2(sub_str, "numericId:") + len("numericId:")
                # index2[3] = find_index2(sub_str, "id=")
                index2[4] = find_index2(sub_str, "numericId=") + len("numericId=")
                index2[5] = find_index2(sub_str, "numeric id") + len("numeric id")
                index2[6] = find_index2(sub_str, "number ") + len("number ")
                index2[7] = find_index2(sub_str, "#") + 1
                index2[8] = find_index2(sub_str, "numericId ") + len("numericId ")
                index2[9] = find_index2(sub_str, "NumericId=") + len("NumericId=")
                index2[10] = find_index2(sub_str, "NumericId ") + len("NumericId ")
                min_index = min(index2)
                # print(min_index)
                sub_str2 = sub_str[min_index:]
                print(sub_str2)
                re=''
                for i in range (0,len(sub_str2)):
                    if sub_str2[i].isdigit():
                        while sub_str2[i].isdigit():
                            re+=sub_str2[i]
                            i+=1
                            if i==len(sub_str2):
                                break
                        break
                # for c in sub_str2:
                #     if c.isdigit():
                #         re+=c
                print(re)
                print("------------")
                # sub_index = sub_str.find("elementId=", sub_str.find("elementId=") + len("elementId="))
                # if sub_index != -1:  # find the answer id
                #     answerId = sub_str[sub_index:]
                # else:
                #     sub_index = sub_str.find("is ", sub_str.find("is ") + len("is "))
                #     if sub_index != -1:  # find the answer id
                #         answerId = sub_str[sub_index:]
                #     else:
                #         # print("didn't find the answer's id\r\n"+data+"----------------")
                #         # print("not found")
                #         continue

                # store in excel
                # Open the workbook and select the sheet


                sheet.cell(row=row_num,column=1,value=name)
                # row_num+=1
                sheet.cell(row=row_num,column=2,value=re)
                row_num+=1
                # Save the changes to the workbook
                workbook.save(output)
def parseMatchResult(data):
    try:
        index1=data.index(".")
        sub_str=data[:index1]
        # print(index1)
        # print(sub_str)
        index2=data.index("numericId: ")
        # index2=find_index(data,"numericId: ")
        sub_str2 = sub_str[index2:]
        # print(sub_str2)
    except ValueError:
        return ''
    re = ''
    for i in range(0, len(sub_str2)):
        if sub_str2[i].isdigit():
            while sub_str2[i].isdigit():
                re += sub_str2[i]
                i += 1
                if i == len(sub_str2):
                    break
            break
    print("parsed selected numericId: "+str(re))
    return re
def write_answer_id(path,output):
    print("1")
    print(path)
    print(output)
    # workbook = openpyxl.load_workbook(output)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "file_name"
    sheet['B1'] = "selected_id"
    sheet['C1'] = "id"
    sheet['D1'] = "name"
    sheet['E1'] = "class"
    sheet['F1'] = "xpath"
    sheet['G1'] = "text"
    sheet['H1'] = "tagName"
    sheet['I1'] = "linkText"
    sheet['J1'] = "location"
    sheet['K1'] = "size"
    sheet['L1'] = "isLeaf"
    sheet['M1'] = "ground_truth_id"

    sheet['O1'] = "id"
    sheet['P1'] = "name"
    sheet['Q1'] = "class"
    sheet['R1'] = "xpath"
    sheet['S1'] = "text"
    sheet['T1'] = "tagName"
    sheet['U1'] = "linkText"
    sheet['V1'] = "location"
    sheet['W1'] = "size"
    sheet['X1'] = "isLeaf"
    sheet['Y1'] = "is_ground_truth"
    sheet['Z1'] = ""
    sheet['AA1'] = "mention"
    sheet['AB1'] = "correct"
    sheet['AC1'] = "correct rate"
    sheet['AD1'] = "answer's index among the four results"
    row_num=2
    print("write id path: "+path)
    for root, dirs, files in os.walk(path):
        for filename in files:
            print(filename)
            if filename.endswith(".txt"):
                filepath = os.path.join(root, filename)
                str_arr=filename.split("_")
                name=str_arr[0]+"_"+str_arr[1]
                # print(filepath)
                with open(filepath, 'r', encoding="utf-8") as f:
                    data = f.read()
                    # print(data)
                # extract answer id
                # index = [100000] * 10
                # index[0] = 100000
                # index[1] = find_index(data, ",")
                # index[2] = find_index(data, ".")
                # index[3] = find_index(data, " because")
                # index[4] = find_index(data, " for")
                # index[5] = find_index(data, " since")
                # index[6] = find_index(data, " based on")
                # index[7] = find_index(data, " according to")
                # index[8] = find_index(data, " as")
                # index[9] = find_index(data, " which")
                #
                # if "Based on" in data:
                #     sub_str = data[index[1]:index[2]]
                # else:
                #     min_index = min(index)
                #     sub_str = data[:min_index]
                # # print(sub_str) # find the substring contains answer id
                # index_is = find_index(sub_str, "is ")
                # print(index_is)
                # if index_is != 10000:
                #     sub_str = sub_str[index_is:]
                #     print(sub_str)
                # else:
                #     index_is = find_index(sub_str, "element's ")
                #     if index_is != 10000:
                #         sub_str = sub_str[index_is:]
                #         print(sub_str)
                # # sub_str = sub_str.replace("{", "")
                # # sub_str = sub_str.replace("}", "")
                # index2 = [10000] * 12
                # index2[0] = find_index2(sub_str, "numericid=") + len("numericid=")
                # index2[1] = find_index2(sub_str, "element ") + len("element ")
                # index2[2] = find_index2(sub_str, "numericId:") + len("numericId:")
                # # index2[3] = find_index2(sub_str, "id=")
                # index2[4] = find_index2(sub_str, "numericId=") + len("numericId=")
                # index2[5] = find_index2(sub_str, "numeric id") + len("numeric id")
                # index2[6] = find_index2(sub_str, "number ") + len("number ")
                # index2[7] = find_index2(sub_str, "#") + 1
                # index2[8] = find_index2(sub_str, "numericId ") + len("numericId ")
                # index2[9] = find_index2(sub_str, "NumericId=") + len("NumericId=")
                # index2[10] = find_index2(sub_str, "NumericId ") + len("NumericId ")
                # min_index = min(index2)
                # # print(min_index)
                # sub_str2 = sub_str[min_index:]
                # print(sub_str2)
                # re=''
                # for i in range (0,len(sub_str2)):
                #     if sub_str2[i].isdigit():
                #         while sub_str2[i].isdigit():
                #             re+=sub_str2[i]
                #             i+=1
                #             if i==len(sub_str2):
                #                 break
                #         break
                re=parseMatchResult(data)
                print(re)
                print("------------")


                sheet.cell(row=row_num,column=1,value=name)
                # row_num+=1
                sheet.cell(row=row_num,column=2,value=re)
                row_num+=1
                # Save the changes to the workbook
    workbook.save(output)

def find_index3(data,str):
    try:
        index=data.index(str)
    except ValueError:
        index=10000
    return index

def find_index(data,str):
    try:
        index=data.index(str)
    except ValueError:
        index=10000
    sub=data[:index]
    if "is" in sub and any(c.isdigit() for c in sub):
        return index
    if "is" in str:
        sub=data[index:]
        if any(c.isdigit() for c in sub):
            return index
    else:
        return 10000
def find_index2(data,str):
    try:
        index=data.index(str)
        # print(index)
    except ValueError:
        index=10001
    # print("index is ")
    # print(index)
    return index

def getGTID(attributes,output,gt_path):
    workbook_attr = openpyxl.load_workbook(attributes)
    worksheet_attr = workbook_attr.active
    workbook_output = openpyxl.load_workbook(output)
    worksheet_output = workbook_output.active

    for i in range(1,worksheet_output.max_row+1):
        gt_id = None
        filename0=worksheet_output.cell(row=i,column=1).value
        selected_id = worksheet_output.cell(row=i,column=2).value
        if filename0==None or "_" not in filename0:
            continue
        filename0 = filename0.replace("_all.txt", "")
        # print("filename0: "+filename0)
        for root, dirs, files in os.walk(gt_path):
            for filename in files:
                # print("filename: "+filename)
                if filename.endswith(".txt") and filename0 in filename:
                    filepath = os.path.join(root, filename)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = f.read()
                    datas = data.split("\n")
                    for content in datas:
                        if "expect xpath:" in content:
                            gt_xpath = content.split(":")[1]
                            print("expect xpath:"+gt_xpath)
                            break
                        else:
                            gt_xpath=datas[0]
        for j in range(1,worksheet_attr.max_row+1):
            filename_att=worksheet_attr.cell(row=j,column=1).value
            if filename_att == None or "_" not in filename_att:
                continue
            # filename=filename.replace(".txt","")
            if filename0 in filename_att:
                for z in range(1,11):
                    c_xpath=worksheet_attr.cell(row=j+z,column=6).value
                    c_id=worksheet_attr.cell(row=j+z,column=2).value
                    if c_id==None:
                        continue
                    c_xpath=c_xpath.replace("'","").replace("[1]","")
                    gt_xpath=gt_xpath.replace("[1]","")
                    print("c_xpath: "+c_xpath)
                    if c_xpath==gt_xpath:
                        print("find gt")
                        worksheet_output.cell(row=i,column=13,value=c_id)
                        gt_id=c_id
                        break
                    else:
                        print(gt_path)
                        print(c_xpath)
                        print()
                break
        if selected_id!=None and gt_id!=None and gt_id==selected_id:
            worksheet_output.cell(row=i, column=25, value=1)
        else:
            worksheet_output.cell(row=i, column=25, value=0)
    workbook_output.save(output)

def getGTID(attributes,output,gt_path):
    workbook_attr = openpyxl.load_workbook(attributes)
    worksheet_attr = workbook_attr.active
    workbook_re = openpyxl.load_workbook(output)
    worksheet_re = workbook_re.active

    for i in range(1,worksheet_re.max_row+1):
        gt_id=None
        filename0=worksheet_re.cell(row=i,column=1).value
        if filename0==None:
            continue
        filename0 = filename0.replace("_all.txt", "")
        # print("filename0: "+filename0)
        name0s=filename0.split("_")
        name0=name0s[0]+"_"+name0s[1]
        # print("name0:"+name0)
        for root, dirs, files in os.walk(gt_path):
            for filename in files:
                # print("filename: "+filename)
                name1s = filename.replace(".txt","").split("_")
                name1 = name1s[0] + "_" + name1s[1]
                # print("name1:"+name1)
                if filename.endswith(".txt") and name0 == name1:
                    filepath = os.path.join(root, filename)
                    # print("open file: "+filepath)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = f.read()
                    datas = data.split("\n")
                    # print(data)
                    for content in datas:
                        if "expect xpath:" in content:
                            gt_xpath = content.split(":")[1]
                            # print("expect xpath:"+gt_xpath)
                            break
                        else:
                            gt_xpath=datas[0]
                            # print("expect xpath:" + gt_xpath)
                            break
                    break
        for j in range(1,worksheet_attr.max_row+1):
            filename_att=worksheet_attr.cell(row=j,column=1).value
            if filename_att == None or "_" not in filename_att:
                continue
            # filename=filename.replace(".txt","")
            # print("filename0: "+filename0+", filename_att:"+filename_att)
            filename_att=filename_att.replace("_all.txt","")
            name2s=filename_att.split("_")
            name2=name2s[0]+"_"+name2s[1]
            if filename0 == name2:
                for z in range(1,11):
                    c_xpath=worksheet_attr.cell(row=j+z,column=6).value
                    c_id=worksheet_attr.cell(row=j+z,column=2).value
                    if c_id==None:
                        continue
                    c_xpath=c_xpath.replace("'","").replace("[1]","")
                    gt_xpath=gt_xpath.replace("[1]","")
                    # print("c_xpath: "+c_xpath)
                    if c_xpath==gt_xpath:
                        # print("find gt")
                        worksheet_re.cell(row=i,column=13,value=c_id)
                        selected_id=worksheet_re.cell(row=i,column=2).value
                        if selected_id!=None:
                            if selected_id==c_id:
                                worksheet_re.cell(row=i, column=25, value=1)
                            else:
                                worksheet_re.cell(row=i, column=25, value=0)
                        break
                    # else:
                    #     print("these two are not the same:")
                    #     print(gt_xpath)
                    #     print(c_xpath)
                    #     print()
                break
    if selected_id != None and gt_id != None and gt_id == selected_id:
        worksheet_re.cell(row=i, column=25, value=1)
    else:
        worksheet_re.cell(row=i, column=25, value=0)
    sum_correct_matching=0
    for i in range(2,worksheet_re.max_row+1):
        is_gt=worksheet_re.cell(row=i,column=25).value
        if is_gt==None:
            is_gt=0
        sum_correct_matching+=int(is_gt)
        if i==5:
            sum_correct_matching+=int(is_gt)*26 #total 27 times of this case
    worksheet_re.cell(row=worksheet_re.max_row+2, column=25, value=sum_correct_matching)
    print(method+" correct matching: "+str(sum_correct_matching))
    workbook_re.save(output)

if __name__ == "__main__":

    gt_paths = "...\\ground_truth\\"
    for i in range(0,4):
        answer_prefix1 = "...\\chatgpt_answer"+str(i)+"\\"
        methods=["vista","water","xpath"]
        result_prefix1 = ""
        attr_prefix1 = "" # need to first generate attributes.xlsx and similarities.xlsx
        for method in methods:
            write_answer_id(answer_prefix1 + method + "\\",result_prefix1+ method + "_match"+str(i)+".xlsx")
            check_without(answer_prefix1 + method + "\\", result_prefix1 + method + "_match"+str(i)+".xlsx")
            attr_path=attr_prefix1+method+"_attributes.xlsx"
            getGTID(attr_path,result_prefix1 + method + "_match"+str(i)+".xlsx",gt_paths)

        # check_without(path_prefix1+method+"\\",result_prefix1+method+"_match1019.xlsx")
    methods = ["vista", "water", "xpath"]


