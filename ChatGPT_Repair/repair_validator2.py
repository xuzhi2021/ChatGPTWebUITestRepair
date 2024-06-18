import openpyxl


def check_relative(path):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.active
    for i in range(2,sheet.max_row):
        strategy1=sheet.cell(row=i,column=3).value
        locator1=sheet.cell(row=i,column=4).value
        strategy2 = sheet.cell(row=i, column=8).value
        locator2 = sheet.cell(row=i, column=9).value
        flag=False
        if strategy1=="xpath" and strategy2=="xpath":
            if "//" in locator1 and "//" in locator2:
                flag=True
        if flag:
            sheet.cell(row=i,column=21,value=1)
        else:
            sheet.cell(row=i,column=21,value=0)
    workbook.save(path)
def check_consistency(attr,output):

    workbook_attr = openpyxl.load_workbook(attr)

    # print("open "+output)
    sheet_attr = workbook_attr.active
    workbook_output=openpyxl.load_workbook(output)
    sheet_output=workbook_output.active
    sheet_output['Z1']="locator same as broken stm"
    num=0
    print(sheet_output.max_row)
    for i in range(2,sheet_output.max_row+1):
        filename=sheet_output.cell(row=i,column=1).value
        selected_id=sheet_output.cell(row=i,column=18).value
        locate_method=sheet_output.cell(row=i,column=8).value
        repaired_content=sheet_output.cell(row=i,column=9).value
        locate_index=get_locator_index(locate_method)
        if locate_index==0:
            sheet_output.cell(row=i,column=19,value=0)
            if locate_method==None:
                print("invalid locator strategy: None, line: "+str(i))
            else:
                print("invalid locator strategy: "+locate_method+", line: "+str(i))
            continue
        print("line "+str(i)+" "+filename)
        for j in range(2,sheet_attr.max_row):
            filename1=sheet_attr.cell(row=j,column=1).value
            if filename1==None:
                continue
            # print("find " + filename + ", current: " + filename1)
            if "-" in filename1:
                continue
            filename1=filename1.replace("_all.txt","")
            # print(filename1)
            found_selected_id_in_attr=False
            if filename == filename1:
                print("found" + filename)
                for z in range(1,11):
                    # print("found"+filename)
                    c_id=sheet_attr.cell(row=j+z,column=2).value
                    if selected_id==None:
                        sheet_output.cell(row=i, column=19, value=0)
                    if c_id==None or selected_id==None:
                        continue
                    if int(c_id)==int(selected_id):
                        found_selected_id_in_attr=True
                        locator_content=sheet_attr.cell(row=j+z,column=locate_index).value
                        locator_content=locator_content.replace("'","")
                        if repaired_content==locator_content:
                            sheet_output.cell(row=i,column=19,value=1)
                        elif locate_method=="xpath" and repaired_content.replace("[1]","")==locator_content.replace("[1]",""):
                            sheet_output.cell(row=i, column=19, value=1)
                        else:
                            print("inconsistent:")
                            print("repaired: "+repaired_content)
                            print("expected: "+locator_content)
                            print("----")
                            sheet_output.cell(row=i,column=19,value=0)
                locator_broken=sheet_output.cell(row=i,column=4).value
                if repaired_content == locator_broken:
                    sheet_output.cell(row=i, column=26, value=1)
                elif locate_method == "xpath" and repaired_content.replace("[1]","") == locator_broken.replace("[1]", ""):
                    sheet_output.cell(row=i, column=26, value=1)
                else:
                    sheet_output.cell(row=i, column=26, value=0)
                # print("finish "+str(num))
                num+=1
                if not found_selected_id_in_attr:
                    sheet_output.cell(row=i, column=19, value=0)
                break
        workbook_output.save(output)


def get_locator_index(sub2):
    locator_index=0
    if sub2==None:
        return 0
    if "xpath" in sub2:
        # index3 = sub2.rindex("By.xpath")+len("By.xpath")
        # locator = "xpath"
        locator_index = 6
    elif "name" in sub2:
        # index3 = sub2.rindex("By.name")+len("By.name")
        # locator = "name"
        locator_index = 4
    elif "linkText" in sub2:
        # index3 = sub2.rindex("By.linkText")+len("By.linkText")
        # locator = "linkText"
        locator_index = 9
    elif "tagName" in sub2:
        # index3 = sub2.rindex("By.tagName")+len("By.tagName")
        # locator = "tagName"
        locator_index = 8
    elif "id" in sub2:
        # index3 = sub2.rindex("By.id")+len("By.id")
        # locator = "id"
        locator_index = 3
    elif "className" in sub2:
        # index3 = sub2.rindex("By.className")+len("By.className")
        # locator = "id"
        locator_index = 5
    return locator_index

def get_locator_index0(sub2):
    locator_index=0
    if sub2==None:
        return 0
    if "xpath" in sub2:
        # index3 = sub2.rindex("By.xpath")+len("By.xpath")
        # locator = "xpath"
        locator_index = 9
    elif "name" in sub2:
        # index3 = sub2.rindex("By.name")+len("By.name")
        # locator = "name"
        locator_index = 4
    elif "linkText" in sub2:
        # index3 = sub2.rindex("By.linkText")+len("By.linkText")
        # locator = "linkText"
        locator_index = 7
    elif "tagName" in sub2:
        # index3 = sub2.rindex("By.tagName")+len("By.tagName")
        # locator = "tagName"
        locator_index = 6
    elif "id" in sub2:
        # index3 = sub2.rindex("By.id")+len("By.id")
        # locator = "id"
        locator_index = 3
    return locator_index

def  write_matching_result(match,output):
    workbook_re = openpyxl.load_workbook(output)
    sheet_re=workbook_re.active
    workbook_match=openpyxl.load_workbook(match)
    sheet_match=workbook_match.active
    print("write matching result to repair")
    for i in range(2,sheet_re.max_row+1):
        filename0=sheet_re.cell(row=i,column=1).value
        print(filename0)
        for j in range(2,sheet_match.max_row+1):
            filename=sheet_match.cell(row=j,column=1).value
            if filename==None:
                continue
            if filename==filename0:
                selected_id=sheet_match.cell(row=j,column=2).value
                sheet_re.cell(row=i,column=18,value=selected_id)
                print(filename+" select "+str(selected_id))
                match_result=sheet_match.cell(row=j,column=25).value
                if match_result==None or match_result=='0' or match_result==0:
                    sheet_re.cell(row=i,column=20,value=0)
                elif match_result=='1' or match_result==1:
                    sheet_re.cell(row=i,column=20,value=1)
    workbook_re.save(output_path)

def check_outside_locator(output):
    workbook=openpyxl.load_workbook(output)
    sheet=workbook.active
    for i in range(2,sheet.max_row+1):
        re1=sheet.cell(row=i,column=12).value
        re4=sheet.cell(row=i,column=15).value
        if re1==1 and re4==1:
            sheet.cell(row=i,column=17,value=1)
        else:
            sheet.cell(row=i,column=17,value=0)
    workbook.save(output)

if __name__ == "__main__":
    methods=["vista","water","xpath"]
    for method in methods:
        for i in range(0,4):
            print(method+"_"+str(i))
            attr_path = ""+method+"_attributes.xlsx"
            output_path = ""+method+"_repair"+str(i)+".xlsx"
            match_path = "" + method + "_match" + str(i) + ".xlsx"
            to_replace = "chatgpt_answerss" + str(i) + "\\" + method
            write_matching_result(match_path, output_path)
            check_outside_locator(output_path)
            check_consistency(attr_path, output_path)  # check the attribute value is equals to the selected element's info
