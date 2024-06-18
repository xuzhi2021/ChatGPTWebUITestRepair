import time
import openai
import os
import openpyxl
import re

def find_file_need_sc(match_path):
    workbook = openpyxl.load_workbook(match_path,data_only=True)
    # print("check correct open "+similarity_path)
    sheet = workbook.active
    filename_list=[]
    for i in range(2,sheet.max_row+1):
        # print("line: "+str(i))
        value=sheet.cell(row=i,column=29).value # ec
        gt=sheet.cell(row=1,column=13).value
        filename=sheet.cell(row=i,column=1).value
        if gt==None or gt=='':
            continue
        filename_list.append(filename)
        # if value==None or filename==None:
        #     continue
        # if float(value)<=0.5: # set thredhold
        #     filename_list.append(filename)
    return filename_list

def ask(last_selection_id, candidates, broken_statement, history_prompt,history_answer, selected_element_info, target, incorrect_attr, sleep_time):
    time.sleep(sleep_time)
    #selection_content,selected_element,selection_time,cur_history,prompt1,sleep_time=askSelection(target, candidates, broken_statement, history,sleep_time) #with history
    #without history
    selection_content, selected_element_info2, cur_selected_id, selection_time, prompt1, sleep_time, invalid = askMatch(history_prompt,history_answer,selected_element_info,candidates,incorrect_attr,sleep_time)
    if invalid or last_selection_id==cur_selected_id:
        #the result may be the same as the last repair, so we don't ask ChatGPT to repair again
        store_answers = selection_content + "\n" +  "selection time: " + str(selection_time)
        prompts = "Prompt of selection:" + "\n" + str(prompt1)
        if invalid:
            print("return before repair because selection answer is invalid")
        else:
            print("the selected element is the same as before")
        return store_answers, selected_element_info2, prompts, sleep_time
    time.sleep(sleep_time)
    repair_content,repair_time,prompt2,sleep_time=askRepair(selected_element_info2,broken_statement,sleep_time)
    store_answers=selection_content+"\n"+repair_content+"\n"+"selection time: "+\
                  str(selection_time)+"\n"+"repair time: "+str(repair_time)
    prompts="Prompt of selection:"+"\n"+str(prompt1)+"\n\n"+"Prompt of repair:"+"\n"+str(prompt2)
    return store_answers,selected_element_info,prompts,sleep_time


def getAnswerInMatchTable(filename,match_xlsx):
    workbook = openpyxl.load_workbook(match_xlsx)
    sheet = workbook.active
    mentioned = []
    correct = []
    incorrect = []
    selectedId=''
    filenames=filename.split("_")
    filename1=filenames[0]+"_"+filenames[1]
    for i in range(2, sheet.max_row + 1):
        filename0 = sheet.cell(row=i, column=1).value
        if filename0 == filename1:
            selectedId = sheet.cell(row=i,column=2).value
            for j in range(3, 13):
                value = sheet.cell(row=i, column=j).value
                if value == None or value == 0:
                    continue
                column_name = sheet.cell(row=1, column=j).value
                mentioned.append(column_name)
            # mentioned=mentioned[0:len(mentioned)-3]
            print(mentioned)
            incorrect = mentioned.copy()
            for j in range(15, 25):
                value = sheet.cell(row=i, column=j).value
                if value == None or value == 0:
                    continue
                column_name = sheet.cell(row=1, column=j).value
                # print("column "+str(j)+" is "+column_name)
                correct.append(column_name)
                incorrect.remove(column_name)
            print(mentioned)
            print(correct)
            print(incorrect)
            mentioned_attr = ''
            for attribute in mentioned:
                mentioned_attr += attribute + ", "
            mentioned_attr = mentioned_attr[0:len(mentioned_attr) - 2]
            print("mention attributes:")
            print(mentioned_attr)
            incorrect_attr = ''
            for attribute in incorrect:
                incorrect_attr += attribute + ", "
            incorrect_attr = incorrect_attr[0:len(incorrect_attr) - 2]
            print(incorrect_attr)

    answer="The most similar element's numericId: "+ selectedId + \
           ". Because they share the most similar attributes:"+mentioned_attr+". "
    return selectedId,answer,incorrect_attr

def process_file(path,to_replace,method,sleep_time,match_xlsx, file_list, index):

    workbook = openpyxl.load_workbook(match_xlsx)
    sheet=workbook.active
    print(path)
    for root, dirs, files in os.walk(path):
        for filename in files:

            if filename.endswith(".txt") and filename.replace("_prompt.txt","") in file_list:
                print("self correct: "+filename)
                filepath = os.path.join(root, filename)#prompt path
                with open(filepath, 'r',encoding="UTF-8") as f:
                    data = f.read()
                target, candidates=getElementInfo(data)
                index1=data.index("Prompt of selection:")+len("Prompt of selection:")
                index2=data.index("Prompt of repair:")
                history_prompt=data[index1:index2] # no difference within different times
                skip=False
                for j in range(2,sheet.max_row+1):
                    filename0=sheet.cell(row=j,column=1).value
                    if filename0==None:
                        continue
                    elif filename.replace("_prompt.txt","")==filename0:
                        ec=sheet.cell(row=j,column=29).value
                        if ec=="1":
                            print("The explanation of " + filename0 + " are consistent with the selection.")
                            skip=True
                            break
                if skip:
                    print("Skip")
                    continue
                print(filename)
                last_selection_id, history_answer, incorrect_attr = getAnswerInMatchTable(filename,match_xlsx)
                # last_selection=parseMatchResult(answer_data)#id
                selected_element_info=getElementById(candidates, last_selection_id)
                print(candidates)
                to_replace0=to_replace+"/"+method
                broken_stm=getBrokenStatement(to_replace0,filepath)
                new_filename = filename.replace("prompt.txt", "answer" + ".txt")
                # for i in range(0,2):
                i=index
                new_path = os.path.join(root.replace(to_replace, to_replace.replace("_prompt","") + "_sc_answers_"+str(i)), new_filename)
                prompt_filename = filename#.replace("all.txt", "prompt" + ".txt")
                prompt_path = os.path.join(root.replace(to_replace, to_replace.replace("_prompt","") + "_sc_prompts_"+str(i)), prompt_filename)
                # print(new_path)
                # print("last_selection: " + last_selection)
                # print("candidates: "+candidates)
                # print(broken_stm)
                print("-------")
                if os.path.exists(new_path):
                    print(new_path + " already exists")
                    continue
                current_timestamp = time.time()
                local_time = time.localtime(current_timestamp)
                formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", local_time)
                print("Current Time:", formatted_time)
                content_store,selected_element,prompts,sleep_time=ask(last_selection_id,candidates,broken_stm,history_prompt,history_answer,selected_element_info,target, incorrect_attr,sleep_time)

                # content_store=selected_id+"\r\n"+explanation+"\r\n"+repaired_Stm+"\r\n"+"runtime:"+str(runtime)+"\r\n"
                if len(candidates)==0:
                    print("no candidate for matching, skip")
                    continue
                # sc_answer,consumed_time,sc_prompt,sleep_time=ask(history_prompt,history_answer,sleep_time)
                os.makedirs(os.path.dirname(new_path), exist_ok=True)
                with open(new_path, 'w', encoding="UTF-8") as f:
                    f.write(content_store)

                os.makedirs(os.path.dirname(prompt_path), exist_ok=True)
                with open(prompt_path, 'w', encoding="UTF-8") as f:
                    f.write(str(prompts))
                print("store sc prompt and answer in: ")
                print(new_path)
                print(prompt_path)
                print("finish one self-correction")
                print("------")
                print()


                content_store = ''



def askMatch(history_prompt,history_answer,selected_element_info,candidates,incorrect_attr,sleep_time):
    print("asking ChatGPT to self correct")
    reply=''

    history = "This is a previous prompt: "+history_prompt+"\n"\
              +"This is your previous answer: "+history_answer+"\n"
    instruction0 = "But your explanation for attributes "+incorrect_attr+" are inconsistent with your selection " \
                    "and this will influce the correctness of your answer. " \
                    "Please answer again." #"Do you want to make another selection?"


    instruction1 = "Your answer should follow the format of this example:"\
                   "\"The most similar element’s numericId: 1. Because they share the most similar attributes: id, xpath, text.\""


    messages3 =  [
        {"role": "system", "content": history+instruction0+instruction1},
        # {"role": "user", "content": user_prompt},
        {"role": "assistant", "content": ''}
    ]
    count=0
    invalid=False
    while reply=='':
        try:
            start1 = time.time()
            reply=openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages3,
                # max_token=token_limit,
                temperature=0.8 #control the stability of answer
            )
            match_result = reply.choices[0].message["content"]
            print(match_result)
            cur_selected_id = parseMatchResult(match_result)
            print("parse current selected id")
            print(cur_selected_id)
            valid_result = True
            for i in cur_selected_id:
                if not i.isdigit():
                    valid_result = False
                    break
            # if not valid_result:
            #     continue  # ask chatgpt again
            # candidates=getElementInfo(history_prompt)
            selected_element = getElementById(candidates, cur_selected_id)
            print("selected element's info")
            print(selected_element)
            if selected_element == '':
                print("parse selected element's id is None")
                valid_result = False
            if not valid_result:
                print(match_result)
                print("parse selected numericId: "+cur_selected_id)
                print("parse selected element's info: "+selected_element)
                print("Get invalid result from ChatGPT, selection is not from candidates. Ask again.")
                time.sleep(sleep_time)
                count+=1
                if count==3:
                    invalid=True
                    end1=time.time()
                    return reply.choices[0].message["content"],'','',end1-start1,messages3,sleep_time,invalid
                continue
            else:
                break
        except openai.error.InvalidRequestError as error:
            print(error)
            print("-----")
            print(messages3)
            print("-----")
            print("This model's maximum context length is 4096 tokens. However, your messages exceeds the limitaion. Please reduce the length of the messages.")
            return
        except openai.error.RateLimitError as error:
            print(f"Rate limit error: {error}")
            print("Waiting for 20 seconds before retrying...")
            if sleep_time<20:
                sleep_time=20
                time.sleep(sleep_time)
            else:
                sleep_time=433
                time.sleep(sleep_time)  # wait for 450 seconds before retrying
        except openai.error.ServiceUnavailableError as error:
            print("openai.error.ServiceUnavailableError: The server is overloaded or not ready yet.")
            time.sleep(sleep_time)

    end1=time.time()
    for choice in reply.choices:
        content = choice.message["content"]
        print(content)

    print("-----")
    print(reply.choices[0].message["content"])
    print("runtime:\r\n"+str(end1-start1)+"\r\n"+"------------------------------------------------")

    return reply.choices[0].message["content"],selected_element,cur_selected_id,end1-start1,messages3,sleep_time,invalid

def askRepair(selected_e, broken_statement,sleep_time):
    reply=''

    history = "You are a web UI test script repair tool. " \
                   "To repair the broken statement, you chose the element "+selected_e+" as the most similar to the target element from the given candidate element list. "
    instruction2="Now based on your selected element, update the locator of the broken statement. Give the result of repaired statement."

    user_prompt2="Broken statement: "+broken_statement

    # token_limit=countToken(instruction0+user_prompt1+instruction2+user_prompt2)

    messages3 =  [
        {"role": "system", "content": history+instruction2},
        {"role": "user", "content": user_prompt2},
        {"role": "assistant", "content": ''}
    ]
    while reply=='':
        try:
            start1 = time.time()
            reply=openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages3,
                # max_token=token_limit,
                temperature=0.8 #control the stability of answer
            )
        except openai.error.InvalidRequestError as error:
            print(error)
            print("-----")
            print(messages3)
            print("-----")
            print("This model's maximum context length is 4096 tokens. However, your messages exceeds the limitaion. Please reduce the length of the messages.")
            return
        except openai.error.RateLimitError as error:
            print(f"Rate limit error: {error}")
            print("Waiting for 20 seconds before retrying...")
            if sleep_time<20:
                sleep_time=20
                time.sleep(sleep_time)
            else:
                sleep_time=433
                time.sleep(sleep_time)  # wait for 450 seconds before retrying
        except openai.error.ServiceUnavailableError as error:
            print("openai.error.ServiceUnavailableError: The server is overloaded or not ready yet.")
            time.sleep(sleep_time)

    end1=time.time()
    for choice in reply.choices:
        content = choice.message["content"]
        print(content)

    print("-----")
    print(reply.choices[0].message["content"])
    print("runtime:\r\n"+str(end1-start1)+"\r\n"+"------------------------------------------------")

    return reply.choices[0].message["content"],end1-start1,messages3,sleep_time


def parseMatchResult0(data):
    numeric_id_match = re.search(r'numericId: (\d+)', data)
    if numeric_id_match:
        numeric_id = numeric_id_match.group(1)
        print("selected numericId:", numeric_id)
        return numeric_id
    else:
        print("selected numericId not found.")
        return ''

def parseMatchResult(data):
    try:
        index1=data.index(".")
        sub_str=data[:index1]
        # print(index1)
        # print(sub_str)
        index2=data.index("numericId: ")
        # index2=find_index(data,"numericId: ")
        sub_str2 = sub_str[index2:]
        # print(sub_str2)
    except ValueError:
        return ''
    re = ''
    for i in range(0, len(sub_str2)):
        if sub_str2[i].isdigit():
            while sub_str2[i].isdigit():
                re += sub_str2[i]
                i += 1
                if i == len(sub_str2):
                    break
            break
    print("parsed selected numericId: "+str(re))
    return re


def find_index(data,str):
    try:
        index=data.index(str)
    except ValueError:
        index=10000
    sub=data[:index]
    if "is" in sub and any(c.isdigit() for c in sub):
        return index
    if "is" in str:
        sub=data[index:]
        if any(c.isdigit() for c in sub):
            return index
    else:
        return 10000

def find_index2(data,str):
    try:
        index=data.index(str)
        # print(index)
    except ValueError:
        index=10001
    # print("index is ")
    # print(index)
    return index

def getSelectedElement(id,candidates):
    elements=candidates.split("}")
    target_str="numericId="+str(id)+","
    re=''
    for e in elements:
        if target_str in e:
            re=e
            break
    return e


def getBrokenStatement(to_replace,filepath):
    # p1 = filepath.replace("chatgpt10_0520/xpath", "broken_statement")
    print("to replace: "+ to_replace)
    print(filepath)
    p1 = filepath.replace(to_replace, "broken_statement")
    p2 = p1.replace("/", "\\")
    p2 = p2.replace("_prompt.txt", ".txt")
    print(p1)
    print(p2)
    with open(p2, 'r', encoding="UTF-8") as f:
        stm = f.read()
    print("broken statement is: "+stm)
    return stm

def getElementInfo(data):
    prefix1="Target element: "
    index1=data.find(prefix1)+len(prefix1)
    prefix2="Candidate elements: "
    index2=data.find(prefix2)+len(prefix2)
    target=data[index1:index2-len(prefix2)]
    prefix3="Prompt of repair"
    index3 = data.find(prefix3) + len(prefix3)
    candidates=data[index2:index3]
    candidates=candidates.replace(prefix3,'')
    candidates=candidates.replace("PreDomNodeInfo{, ","{")
    return target,candidates

def getElementById(candidates,id):
    # print("------")
    # print(candidates)
    # print("------")
    elements = re.findall(r'\{.*?\}', candidates)
    target_str="numericId="+str(id)
    result=''
    for e in elements:
        if target_str in e:
            result=e
            break
    return result

def getIncorrectRepairs(union_path):
    workbook=openpyxl.load_workbook(union_path)
    sheet=workbook.active
    file_list=[]
    for i in range(2,sheet.max_row+1):
        filename=sheet.cell(row=i,column=1).value
        if filename==None:
            continue
        correct=sheet.cell(row=i,column=6).value
        if int(correct)==0:
            file_list.append(filename)
    return file_list

if __name__ == "__main__":
    openai.api_key = ""
    to_replace="chatgpt_prompts"
    methods=["xpath","water","vista"]
    webs=["addressbook","Claroline_Test_Suite","collabtive","mantisbt","mrbs"]
    for method in methods:
        for web in webs:
            path_test = ".../chatgpt_prompts/" \
                    + method+"/testcases/"+web+"/"
            match_path=""+method+"_match.xlsx"

    for method in methods:
        for i in range(0,4):
            original_data_path = ".../chatgpt10_1109"+"/"+method
            match_path = "..." + method + "_match"+str(i)+".xlsx"
            path_test = "/chatgpt10_prompts"+str(i)+"/"+method

            to_replace="chatgpt_prompts"+str(i)
            repair_result_path=""+method+"_result.xlsx"
            file_list=getIncorrectRepairs(repair_result_path)
            print(method)
            print(file_list)
            print(len(file_list))
            process_file(path_test,to_replace,method,1,match_path,file_list,i)